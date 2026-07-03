"""Normalize uploaded or pasted test pack content into text."""

from __future__ import annotations

import csv
import io
from pathlib import Path


def ingest_text(content: str) -> str:
    return (content or "").strip()


def ingest_file(path: str | Path, *, filename: str | None = None) -> str:
    path = Path(path)
    name = (filename or path.name).lower()

    if name.endswith((".md", ".txt")):
        return path.read_text(encoding="utf-8", errors="ignore")
    if name.endswith(".csv"):
        return _csv_to_text(path.read_text(encoding="utf-8", errors="ignore"))
    if name.endswith((".xlsx", ".xls")):
        return _xlsx_to_text(path)
    return path.read_text(encoding="utf-8", errors="ignore")


def _csv_to_text(raw: str) -> str:
    reader = csv.reader(io.StringIO(raw))
    lines = [" | ".join(row) for row in reader if any(cell.strip() for cell in row)]
    return "\n".join(lines)


def _xlsx_to_text(path: Path) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl required for Excel uploads: pip install openpyxl") from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    parts: list[str] = []
    for sheet in wb.worksheets:
        parts.append(f"## Sheet: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if any(cells):
                parts.append(" | ".join(cells))
    wb.close()
    return "\n".join(parts)
